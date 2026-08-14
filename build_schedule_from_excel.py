"""Optional: rebuild schedule.json from the original local Excel files
instead of the live Google Sheet. Uses the same parsing.py logic as
sheets_fetch.py so behavior always stays in sync between both paths.
"""
import json
import openpyxl

from parsing import build_schedule, clean_date

BFS_FILE = "/mnt/user-data/uploads/PGDMBFS_2025-2027_TERM_IVSchedule.xlsx"
COMBINED_FILE = "/mnt/user-data/uploads/Term_IV_Class_Schedule_2025-27.xlsx"


def extract_rows(path, max_row):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for r in range(5, max_row + 1, 4):
        date_val = ws.cell(r, 1).value
        day_val = ws.cell(r, 2).value
        if date_val is None and day_val is None:
            continue
        d = clean_date(date_val)
        if not d:
            continue
        sessions = {}
        for sname, c in [('S1', 3), ('S2', 4), ('S3', 5), ('S4', 7), ('S5', 8), ('S6', 9)]:
            sessions[sname] = ws.cell(r, c).value
        rows.append({'date': d, 'day': day_val, 'sessions': sessions})
    return rows


if __name__ == "__main__":
    combined_rows = extract_rows(COMBINED_FILE, 424)
    bfs_rows = extract_rows(BFS_FILE, 376)

    schedule = build_schedule([
        (combined_rows, "combined"),
        (bfs_rows, "bfs"),
    ])

    with open("schedule.json", "w") as f:
        json.dump(schedule, f, indent=2)

    print(f"Rebuilt schedule.json — {len(schedule)} dates.")
